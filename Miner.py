from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
import time






def webdata(inp):
    li=[]
    pr=[]
    liuk=[]
    pruk=[]
    driver = webdriver.Chrome()


    driver.get(f'https://www.google.co.uk/search?q={inp}')

    time.sleep(2)

    f=driver.find_elements(By.XPATH,'//div[@class="MjjYud"]')
    for i in f:
        if "£" in i.text:
            link=i.find_element(By.XPATH,'.//a').get_attribute('href')
            price=((i.text).find("€"))
            price=i.text[price-1:price+6]
            print(link,price)
            liuk.append(link)
            pruk.append(price)
    if len(liuk)==0:
        liuk.append("Not Found")
        pruk.append("Not Found")

    time.sleep(5)

    driver.get(f'https://www.ebay.co.uk/sch/i.html?_from=R40&_nkw={inp}&_sacat=0&LH_TitleDesc=0&_fsrp=1&rt=nc&_ipg=240')
    #https://www.ebay.co.uk/sch/i.html?_from=R40&_nkw={inp}&_sacat=0&LH_TitleDesc=0&_fsrp=1&rt=nc&_ipg=240
    time.sleep(3)
    dat=[]

    time.sleep(2)
    data=driver.find_elements(By.XPATH,'//li[@class="s-item s-item__pl-on-bottom"]')
    print(len(data))
    for i in data:
    
        try: 
            f=i.find_element(By.XPATH,'.//span[@class="s-item__location s-item__itemLocation"]')
            print(f.text)
            if 'from United Kingdom' in f.text:
                dat.append(i)
            else:    
                print("Not In")

        except:
            print("In")
            dat.append(i)
    i=0
    print(len(dat))
    while i < len(dat):
        price=dat[i].find_element(By.XPATH,'.//span[@class="s-item__price"]')
        #print(price.text)
        link=dat[i].find_element(By.XPATH,'.//a').get_attribute('href')
        #print(link)
        time.sleep(2)
        li.append(link)
        pr.append(price.text)
        i+=1

    time.sleep(5)


    driver.close()

    return li,pr,liuk,pruk









fn=input("Enter The File Name With Path: ")
on=input("Enter The Output File Name: ")

fn=fn.replace('\\','\\\\').replace("'","\"")
on=on.replace('\\','\\\\').replace("'","\"")
#
# Load the Excel file
wb = load_workbook(filename=fn)

# Select the active sheet
ws = wb.active

# Extract data from column 'C'
col_c = []
for cell in ws['C']:
    col_c.append(cell.value)

# Print the data
print(col_c)
alld=[]


alldat=[]
o=1
col_c = col_c[1:] 

print(col_c)
for i in col_c:
    datp=[]
    datl=[]
    datukp=[]
    datukli=[]
    for k in i.split('\n'):
        alld.append(k)
        print(k,o)
        li,pr,liuk,pruk=webdata(k)
        datl.append(li)
        datp.append(pr)
        datukp.append(pruk)
        datukli.append(liuk)
    dataukprice=str(datukp).replace("'","").replace("[","").replace("]","").replace(", ","\n")
    datauklink=str(datukli).replace("'","").replace("[","").replace("]","").replace(", ","\n")

    dataprice=str(datp).replace("'","").replace("[","").replace("]","").replace(", ","\n")
    datalink=str(datl).replace("'","").replace("[","").replace("]","").replace(", ","\n")

    ws.cell(row=o+1,column=6).value=datauklink
    ws.cell(row=o+1,column=7).value=dataukprice
    ws.cell(row=o+1,column=8).value=datalink
    ws.cell(row=o+1,column=9).value=dataprice
    o+=1

alld=alld[1:]
wb.save(on)
#print(alldat)
